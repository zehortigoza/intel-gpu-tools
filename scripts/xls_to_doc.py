#!/usr/bin/env python3
# pylint: disable=C0301,R0912,R0913,R0914,R0915,R1702
# SPDX-License-Identifier: (GPL-2.0 OR MIT)

## Copyright (C) 2023    Intel Corporation                 ##
## Author: Mauro Carvalho Chehab <mchehab@kernel.org>      ##

"""Import contents of a XLS file into testplan documentation."""

import argparse
import json
import os
import re
import sys

from openpyxl import load_workbook

from test_list import TestList


#
# FillTests class definition
#


class FillTests(TestList):
    """
    Fill documentation source code test comments from a spreadsheet.
    """

    def __init__(self, config_path, verbose):
        self.tests = {}
        self.spreadsheet_data = {}
        self.ignore_fields = []
        self.verbose = verbose

        # Read current documentation
        TestList.__init__(self, config_path)

        self.testname_regex = re.compile(r'^\s*(igt@[^\n\@]+)\@?(\S*)\s*')
        self.key_has_wildcard = re.compile(r'\%?arg\[(\d+)\]')
        self.field_re = re.compile(r"(" + '|'.join(self.field_list.keys()) + r'):\s*(.*)', re.I)

        for test in self.doc:                   # pylint: disable=C0206
            fname = self.doc[test]["File"]

            name = re.sub(r'.*/', '', fname)
            name = re.sub(r'\.[\w+]$', '', name)
            name = "igt@" + name

            subtest_array = self.expand_subtest(fname, name, test, True, True, True)
            for subtest_dict in subtest_array:
                name = subtest_dict["_summary_"]
                del subtest_dict["_summary_"]

                match = self.testname_regex.match(name)
                if not match:
                    sys.exit(f"Error: can't parse {name}")

                testname = match.group(1)
                if match.group(2):
                    subtest = match.group(2)
                else:
                    subtest = ''

                if testname not in self.tests:
                    self.tests[testname] = {}
                    self.tests[testname]["subtests"] = {}

                    self.tests[testname]["Test"] = test
                    self.tests[testname]["File"] = fname

                self.tests[testname]["subtests"][subtest] = subtest_dict

        for field, item in self.props.items():
            if "sublevel" in item["_properties_"]:
                update = item["_properties_"].get("update-from-file")
                if update:
                    self.ignore_fields.append(field)

    def add_field(self, dic, field, value):
        """
        Add a value to a field, alphabetically sorting it.
        """

        if field in dic and dic[field] != '':
            fields = sorted(dic[field].split(", "))
            fields.append(value)
            value = ", ".join(sorted(fields))

        dic[field] = value

    def process_spreadsheet_sheet(self, sheet):

        """
        Convert a single sheet into a dictionary.

        Please notice that an empty row means to ignore the rest of the
        sheet.
        """

        column_list = []
        for cell in sheet[1]:
            column_list.append(cell.value)

        for row in range(2, sheet.max_row):
            if sheet[row][0].value is None:
                print(f"Ignoring sheet after A{row} row, as test name is empty")
                return
            if not isinstance(sheet[row][0].value, str):
                print(f"Ignoring A{row} row on {sheet.title}: test name is not a string: {sheet[row][0].value}")
                continue
            test_name = sheet[row][0].value.strip()
            if not re.match(r'^igt\@', test_name):
                print(f"Ignoring A{row} row on {sheet.title}: not a valid test name: {test_name}")
                continue

            if test_name not in self.spreadsheet_data:
                self.spreadsheet_data[test_name] = {}

            i = 1
            for col in range(2, sheet.max_column + 1):
                val = sheet.cell(row=row, column=col).value
                if val:
                    if isinstance(val, str):
                        val = val.strip()

                    self.spreadsheet_data[test_name][column_list[i]] = val

                i += 1

    def read_spreadsheet_file(self, fname, sheets):
        """
        Read a XLS file, processing all sheets and returing a dict with
        all lines from all sheets altogether.
        """

        # Iterate the loop to read the cell values
        wb = load_workbook(filename=fname)

        # Handle first "normal" sheets
        for sheet in wb:
            if sheets and sheet.title not in sheets:
                continue

            self.process_spreadsheet_sheet(sheet)

        return self.spreadsheet_data

    def change_value(self, content, subtest, line, field, value):
        """
        Change the contents of a source file to update its documentation.
        """

        current_field = None
        i = line
        while True:
            i += 1
            if i >= len(content):
                break

            file_line = content[i]

            if re.match(r'^\s*\*\/\s*$', file_line):
                break

            file_line = re.sub(r'^\s*\* ?', '', file_line)

            match = re.match(r'^SUBTESTS?:\s*(.*)', file_line)
            if match and match.group(1) != subtest:
                break

            match = re.match(r'^TEST:\s*(.*)', file_line)
            if match and match.group(1) != subtest:
                break

            match = re.match(r'arg\[(\d+)\]:\s*(.*)', file_line)
            if match:
                break

            match = re.match(r'\@(\S+):\s*(.*)', file_line)
            if match:
                break

            match = re.match(r'arg\[(\d+)\]\.values:\s*(.*)', file_line)
            if match:
                break

            match = re.match(self.field_re, file_line)
            if match:
                current_field = self.field_list[match.group(1).lower()]
                if current_field != field:
                    continue
                content[i] = ""

            # Handle continuation lines
            if current_field:
                match = re.match(r'\s+(.*)', file_line)
                if match:
                    if current_field != field:
                        continue

                    content[i] = ""

        content.insert(i, f' * {field}: {value}\n')

    def parse_spreadsheet(self, fname, sheets=None):
        """
        Convert a spreadsheet file into a recursive dictionary.

        It assumes that the first column at the parsed sheets will contain
        an IGT test name in the format:

            igt@<testname>@<subtest>

        And the other columns will have the field name at the first line.

        On success, it will return a dictionary with this format:

           self.tests[testname]["subtests"][subtest][field] = value
        """

        if not os.path.isfile(fname):
            print(f'Warning: {fname} not found. Skipping spreadsheet parser')
            return

        data = self.read_spreadsheet_file(fname, sheets)

        for test, row in sorted(data.items()):
            match = self.testname_regex.match(test)
            if not match:
                sys.exit(f"Error: can't parse {test}")

            testname = match.group(1)
            if match.group(2):
                subtest = match.group(2)
            else:
                subtest = ''

            if testname not in self.tests:
                print(f"Warning: {testname} file is not present at JSON config file.")
                continue

            if subtest not in self.tests[testname]["subtests"]:
                self.tests[testname]["subtests"][subtest] = {}

            for key, value in row.items():
                self.tests[testname]["subtests"][subtest][key] = value

    def update_test_file(self, testname, args):
        """
        Update a C source file using the contents of self.tests as
        the source of data to be filled at the igt_doc documentation
        comments.
        """

        try:
            sourcename = self.tests[testname]["File"]
            with open(sourcename, 'r', encoding='utf8') as in_fp:
                content = in_fp.read().splitlines(True)
        except EnvironmentError:
            sys.exit(f'Failed to read {sourcename}')

        try:

            test_nr = self.tests[testname]["Test"]

            for subtest, subtest_content in sorted(self.tests[testname]["subtests"].items()):
                if "line" not in subtest_content:
                    print(f"Warning: didn't find where {subtest} is documented.")
                    continue

                line = subtest_content['line']
                subtest_nr = subtest_content['subtest_nr']

                if subtest_nr not in self.doc[test_nr]["subtest"]:
                    if self.verbose:
                        print(f"Error: missing subtest {subtest_nr} at {self.doc[test_nr]['subtest']}")
                    else:
                        print(f"Warning: test {testname}, subtest {subtest} is not documented.")
                    continue

                doc_content = self.doc[test_nr]["subtest"][subtest_nr]

                # Handling wildcards is not easy. Let's just skip those
                for field, value in sorted(subtest_content.items()):
                    if field in ['line', 'subtest_nr']:
                        continue

                    if args.ignore_lists:
                        if field in self.ignore_fields:
                            continue

                    doc_value = doc_content.get(field)
                    if doc_value:
                        if self.key_has_wildcard.search(doc_value):
                            print(f"Warning: {subtest} field {field} has wildcards. Skipping it.")
                            continue
                        if doc_value == value:
                            if self.verbose > 1:
                                print(f"{testname}@{subtest} field {field}: Value unchanged. Ignoring it")
                            continue

                    if self.verbose > 0:
                        print(f"Update {testname}@{subtest} field {field} on line {line}:")
                        print(f"  Change from {doc_value} to {value}")

                    # Just in case, handle continuation lines
                    value = re.sub(r"\n", "\n *   ", value)

                    self.change_value(content, subtest, line, field, value)

                    # Update line numbers after insert
                    skip = True
                    for sub, sub_content in sorted(self.tests[testname]["subtests"].items()):
                        if sub == subtest:
                            skip = False
                            continue
                        if skip:
                            continue
                        sub_line = sub_content['line']
                        if sub_line >= line:
                            sub_content['line'] += 1

        except EnvironmentError as err:
            sys.exit(f'Error: {err}')

        # Write changes
        try:
            if self.verbose:
                print(f"Writing to {sourcename}")

            with open(sourcename, 'w', encoding='utf8') as out_fp:
                out_fp.write("".join(content))
        except EnvironmentError:
            print(f'Failed to write to {sourcename}')

    def update_test_files(self, args):
        """
        Populate all test files with the documentation from self.tests.
        """

        if self.verbose == 0:
            print("Update source files")

        for testname in self.tests:
            self.update_test_file(testname, args)

######
# Main
######


def main():
    """Main program"""

    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--config", required=True,
                        help="JSON file describing the test plan template")
    parser.add_argument("--xls", required=True,
                        help="Input XLS file.")
    parser.add_argument("--sheets", nargs="*",
                        help="Input only some specific sheets from the XLS file.")
    parser.add_argument('--ignore-lists', action='store_false', default=True,
                        help='Ignore fields that are updated via test lists')
    parser.add_argument("--store-json", action="store_true",
                        help="Generate JSON files with documentation. Useful for debugging purposes.")
    parser.add_argument('-v', '--verbose', action='count', default=0)

    parse_args = parser.parse_args()

    fill_test = FillTests(parse_args.config, parse_args.verbose)

    fill_test.parse_spreadsheet(parse_args.xls, parse_args.sheets)

    if "store_json" in parse_args:
        print("Generating fill_test.json debug file")
        with open("fill_test.json", "w", encoding='utf8') as write_file:
            json.dump(fill_test.tests, write_file, indent=4)

        print("Generating doc.json debug file")
        with open("doc.json", "w", encoding='utf8') as write_file:
            json.dump(fill_test.doc, write_file, indent=4)

    fill_test.update_test_files(parse_args)


if __name__ == '__main__':
    main()
